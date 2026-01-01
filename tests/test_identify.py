"""Tests for file identification (identify method).

The identify() method determines if a file should be processed by this importer.
It checks:
1. File extension (.xlsx)
2. Excel headers match expected DNB format
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from beancount_no_dnb.mastercard import DnbMastercardConfig, Importer


class TestIdentifyBasics:
    """Basic file identification tests."""

    def test_identifies_valid_excel_file(self, basic_importer, sample_excel_path):
        """Correctly identifies a valid DNB Excel file."""
        if sample_excel_path.exists():
            assert basic_importer.identify(str(sample_excel_path)) is True

    def test_identifies_minimal_excel_file(self, basic_importer, minimal_excel_file):
        """Identifies a minimal valid Excel file."""
        assert basic_importer.identify(str(minimal_excel_file)) is True

    def test_rejects_wrong_extension(self, basic_importer, tmp_path):
        """Rejects files without .xlsx extension."""
        txt_file = tmp_path / "statement.txt"
        txt_file.write_text("not an excel file")
        assert basic_importer.identify(str(txt_file)) is False

    def test_rejects_csv_file(self, basic_importer, tmp_path):
        """Rejects CSV files."""
        csv_file = tmp_path / "statement.csv"
        csv_file.write_text("Dato,Beløpet gjelder,Valuta,Kurs,Inn,Ut")
        assert basic_importer.identify(str(csv_file)) is False

    def test_rejects_wrong_headers(self, basic_importer, tmp_path):
        """Rejects Excel files with wrong headers."""
        wb = Workbook()
        ws = wb.active
        # Wrong headers
        headers = ["Date", "Description", "Amount"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        file_path = tmp_path / "wrong_headers.xlsx"
        wb.save(file_path)

        assert basic_importer.identify(str(file_path)) is False


class TestIdentifyEdgeCases:
    """Edge cases for file identification."""

    def test_nonexistent_file(self, basic_importer, tmp_path):
        """Gracefully handles nonexistent files."""
        result = basic_importer.identify(str(tmp_path / "nonexistent.xlsx"))
        assert result is False

    def test_empty_excel_file(self, basic_importer, tmp_path):
        """Handles empty Excel files gracefully."""
        wb = Workbook()
        file_path = tmp_path / "empty.xlsx"
        wb.save(file_path)

        result = basic_importer.identify(str(file_path))
        assert result is False  # No headers

    def test_corrupted_file(self, basic_importer, tmp_path):
        """Handles corrupted files gracefully."""
        garbage_file = tmp_path / "corrupted.xlsx"
        garbage_file.write_bytes(b"\x00\x01\x02\x03\xff\xfe")
        result = basic_importer.identify(str(garbage_file))
        assert result is False


class TestAccountMethod:
    """Tests for the account() method."""

    def test_returns_configured_account_name(self, basic_importer, minimal_excel_file):
        """account() returns the configured account name."""
        result = basic_importer.account(str(minimal_excel_file))
        assert result == "Liabilities:CreditCard:DNB"

    def test_different_account_configurations(self, minimal_excel_file):
        """Different configs return different account names."""
        configs = [
            "Liabilities:CreditCard:DNB:Personal",
            "Liabilities:CreditCard:DNB:Business",
        ]

        for account_name in configs:
            config = DnbMastercardConfig(
                account_name=account_name,
                currency="NOK",
            )
            importer = Importer(config=config, debug=False)
            assert importer.account(str(minimal_excel_file)) == account_name


class TestFilenameMethod:
    """Tests for the filename() method."""

    def test_filename_includes_account_suffix(self, basic_importer, minimal_excel_file):
        """Filename includes the account name suffix."""
        result = basic_importer.filename(str(minimal_excel_file))
        assert result == "dnb_mastercard.DNB.minimal.xlsx"

    def test_filename_preserves_original_basename(self, basic_importer, tmp_path):
        """Original filename is preserved in the result."""
        wb = Workbook()
        ws = wb.active
        headers = ["Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        file_path = tmp_path / "DNB_Statement_2025-12.xlsx"
        wb.save(file_path)

        result = basic_importer.filename(str(file_path))
        assert "DNB_Statement_2025-12.xlsx" in result
