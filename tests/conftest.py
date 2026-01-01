"""Shared pytest fixtures for beancount-no-dnb tests."""

import datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from beancount_no_dnb.mastercard import DnbMastercardConfig, Importer
from beancount_classifier import TransactionPattern, amount
from beancount_no_dnb.models import (
    ExcelFileData,
    ParsedTransaction,
    RawTransaction,
)


# =============================================================================
# Path Fixtures
# =============================================================================


@pytest.fixture
def test_data_dir() -> Path:
    """Path to the test_data directory containing sample Excel files."""
    return Path(__file__).parent.parent / "test_data"


@pytest.fixture
def sample_excel_path(test_data_dir) -> Path:
    """Path to the sample Excel file."""
    return test_data_dir / "sample_statement.xlsx"


# =============================================================================
# Raw Data Fixtures
# =============================================================================


@pytest.fixture
def raw_transaction_debit() -> RawTransaction:
    """A typical debit transaction."""
    return RawTransaction(
        date=datetime.date(2025, 10, 24),
        description="REMA 1000 OSLO, Oslo",
        foreign_currency=None,
        exchange_rate=None,
        credit=None,
        debit=Decimal("150.50"),
    )


@pytest.fixture
def raw_transaction_credit() -> RawTransaction:
    """A credit (refund/payment) transaction."""
    return RawTransaction(
        date=datetime.date(2025, 10, 25),
        description="Innbetaling",
        foreign_currency=None,
        exchange_rate=None,
        credit=Decimal("5000.00"),
        debit=None,
    )


@pytest.fixture
def raw_transaction_balance_forward() -> RawTransaction:
    """A balance forward transaction."""
    return RawTransaction(
        date=datetime.date(2025, 11, 10),
        description="Skyldig beløp fra forrige faktura",
        foreign_currency=None,
        exchange_rate=None,
        credit=None,
        debit=Decimal("5000.00"),
    )


# =============================================================================
# Configuration Fixtures
# =============================================================================


@pytest.fixture
def basic_config() -> DnbMastercardConfig:
    """Basic importer configuration."""
    return DnbMastercardConfig(
        account_name="Liabilities:CreditCard:DNB",
        currency="NOK",
    )


@pytest.fixture
def config_with_mappings() -> DnbMastercardConfig:
    """Configuration with transaction patterns for categorization."""
    return DnbMastercardConfig(
        account_name="Liabilities:CreditCard:DNB",
        currency="NOK",
        transaction_patterns=[
            TransactionPattern(narration="VINMONOPOLET", account="Expenses:Groceries"),
            TransactionPattern(narration="SPOTIFY", account="Expenses:Entertainment:Music"),
            TransactionPattern(narration="REMA", account="Expenses:Groceries"),
            TransactionPattern(narration="GITHUB", account="Expenses:Cloud-Services"),
        ],
    )


@pytest.fixture
def config_include_payments() -> DnbMastercardConfig:
    """Configuration that includes payment entries."""
    return DnbMastercardConfig(
        account_name="Liabilities:CreditCard:DNB",
        currency="NOK",
        skip_payments=False,
        skip_balance_forward=True,
    )


@pytest.fixture
def config_include_all() -> DnbMastercardConfig:
    """Configuration that includes all entries."""
    return DnbMastercardConfig(
        account_name="Liabilities:CreditCard:DNB",
        currency="NOK",
        skip_payments=False,
        skip_balance_forward=False,
    )


# =============================================================================
# Importer Fixtures
# =============================================================================


@pytest.fixture
def basic_importer(basic_config) -> Importer:
    """An importer with basic configuration."""
    return Importer(config=basic_config, debug=False)


@pytest.fixture
def importer_with_mappings(config_with_mappings) -> Importer:
    """An importer configured with categorization mappings."""
    return Importer(config=config_with_mappings, debug=False)


@pytest.fixture
def importer_include_payments(config_include_payments) -> Importer:
    """An importer that includes payment entries."""
    return Importer(config=config_include_payments, debug=False)


@pytest.fixture
def importer_include_all(config_include_all) -> Importer:
    """An importer that includes all entries."""
    return Importer(config=config_include_all, debug=False)


# =============================================================================
# Excel File Content Fixtures
# =============================================================================


@pytest.fixture
def minimal_excel_file(tmp_path) -> Path:
    """Create a minimal valid Excel file for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "transaksjonsliste"

    # Headers
    headers = ["Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Single transaction
    ws.cell(row=2, column=1, value=datetime.datetime(2025, 10, 24))
    ws.cell(row=2, column=2, value="TEST MERCHANT")
    ws.cell(row=2, column=6, value=100.00)

    file_path = tmp_path / "minimal.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def excel_with_all_types(tmp_path) -> Path:
    """Create an Excel file with all transaction types."""
    wb = Workbook()
    ws = wb.active
    ws.title = "transaksjonsliste"

    headers = ["Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    transactions = [
        # Balance forward
        (datetime.datetime(2025, 11, 10), "Skyldig beløp fra forrige faktura", None, None, None, 5000.00),
        # Payment
        (datetime.datetime(2025, 10, 25), "Innbetaling", None, None, 5000.00, None),
        # Debit
        (datetime.datetime(2025, 10, 24), "REMA 1000 OSLO, Oslo", None, None, None, 150.50),
        # Credit/refund
        (datetime.datetime(2025, 10, 29), "Refund - Something", None, None, 50.00, None),
    ]

    for row_num, txn in enumerate(transactions, 2):
        for col, value in enumerate(txn, 1):
            ws.cell(row=row_num, column=col, value=value)

    file_path = tmp_path / "all_types.xlsx"
    wb.save(file_path)
    return file_path
