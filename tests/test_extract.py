"""Integration tests for the full extraction pipeline (extract method).

These tests verify the complete data flow:
    Excel File -> parse -> transform -> categorize -> Beancount Directives
"""

import datetime
from decimal import Decimal

import pytest
from beancount.core import data
from beancount.core.number import D
from openpyxl import Workbook

from beancount_no_dnb.mastercard import DnbMastercardConfig, Importer


class TestExtractBasics:
    """Basic extraction tests."""

    def test_extract_returns_list_of_directives(self, basic_importer, minimal_excel_file):
        """extract() returns a list of Beancount directives."""
        entries = basic_importer.extract(str(minimal_excel_file), [])
        assert isinstance(entries, list)
        assert len(entries) > 0

    def test_extract_creates_transaction_directive(self, basic_importer, minimal_excel_file):
        """Each transaction in the Excel becomes a Transaction directive."""
        entries = basic_importer.extract(str(minimal_excel_file), [])

        transactions = [e for e in entries if isinstance(e, data.Transaction)]
        assert len(transactions) >= 1

    def test_transaction_has_correct_date(self, basic_importer, minimal_excel_file):
        """Transaction date is correctly parsed."""
        entries = basic_importer.extract(str(minimal_excel_file), [])

        transactions = [e for e in entries if isinstance(e, data.Transaction)]
        assert transactions[0].date == datetime.date(2025, 10, 24)

    def test_transaction_has_correct_amount(self, basic_importer, minimal_excel_file):
        """Transaction amount is correctly parsed (debits are negative)."""
        entries = basic_importer.extract(str(minimal_excel_file), [])

        transactions = [e for e in entries if isinstance(e, data.Transaction)]
        txn = transactions[0]
        assert txn.postings[0].units.number == D("-100.00")

    def test_transaction_has_correct_account(self, basic_importer, minimal_excel_file):
        """Transaction uses the configured account name."""
        entries = basic_importer.extract(str(minimal_excel_file), [])

        transactions = [e for e in entries if isinstance(e, data.Transaction)]
        txn = transactions[0]
        assert txn.postings[0].account == "Liabilities:CreditCard:DNB"

    def test_transaction_has_correct_narration(self, basic_importer, minimal_excel_file):
        """Transaction narration is the description from Excel."""
        entries = basic_importer.extract(str(minimal_excel_file), [])

        transactions = [e for e in entries if isinstance(e, data.Transaction)]
        txn = transactions[0]
        assert txn.narration == "TEST MERCHANT"


class TestExtractCreditsAndDebits:
    """Tests for handling credits (Inn) and debits (Ut)."""

    def test_debit_transactions_are_negative(self, basic_importer, excel_with_all_types):
        """Debit transactions (Ut column) have negative amounts."""
        entries = basic_importer.extract(str(excel_with_all_types), [])
        transactions = [e for e in entries if isinstance(e, data.Transaction)]

        # Find debit transaction (REMA)
        rema_txn = next(
            (t for t in transactions if "REMA" in (t.narration or "")), None
        )
        assert rema_txn is not None
        assert rema_txn.postings[0].units.number < 0

    def test_credit_transactions_are_positive(self, basic_importer, excel_with_all_types):
        """Credit transactions (Inn column) have positive amounts."""
        entries = basic_importer.extract(str(excel_with_all_types), [])
        transactions = [e for e in entries if isinstance(e, data.Transaction)]

        # Find credit transaction (Refund)
        refund_txn = next(
            (t for t in transactions if "Refund" in (t.narration or "")), None
        )
        assert refund_txn is not None
        assert refund_txn.postings[0].units.number > 0


class TestExtractSkipBehavior:
    """Tests for skip configuration options."""

    def test_balance_forward_skipped_by_default(self, basic_importer, excel_with_all_types):
        """Balance forward entries are skipped by default."""
        entries = basic_importer.extract(str(excel_with_all_types), [])
        transactions = [e for e in entries if isinstance(e, data.Transaction)]

        # Should not find balance forward
        balance_txn = next(
            (t for t in transactions if "forrige faktura" in (t.narration or "")), None
        )
        assert balance_txn is None

    def test_payments_included_by_default(self, basic_importer, excel_with_all_types):
        """Payment entries are included by default."""
        entries = basic_importer.extract(str(excel_with_all_types), [])
        transactions = [e for e in entries if isinstance(e, data.Transaction)]

        # Should find payment (Innbetaling)
        payment_txn = next(
            (t for t in transactions if "Innbetaling" in (t.narration or "")), None
        )
        assert payment_txn is not None

    def test_include_all_entries(self, importer_include_all, excel_with_all_types):
        """When configured, all entries including balance forward are included."""
        entries = importer_include_all.extract(str(excel_with_all_types), [])
        transactions = [e for e in entries if isinstance(e, data.Transaction)]

        # All 4 transactions should be present
        assert len(transactions) == 4


class TestExtractWithCategorization:
    """Tests for extraction with categorization mappings."""

    def test_matching_transactions_get_categorized(
        self, importer_with_mappings, sample_excel_path
    ):
        """Transactions matching patterns get expense accounts."""
        if not sample_excel_path.exists():
            pytest.skip("Sample Excel file not available")

        entries = importer_with_mappings.extract(str(sample_excel_path), [])
        transactions = [e for e in entries if isinstance(e, data.Transaction)]

        # Find the VINMONOPOLET transaction
        vinmonopolet_txn = next(
            (t for t in transactions if "VINMONOPOLET" in (t.narration or "")), None
        )

        assert vinmonopolet_txn is not None
        # Should have two postings (credit card + expense)
        assert len(vinmonopolet_txn.postings) == 2
        assert vinmonopolet_txn.postings[1].account == "Expenses:Groceries"


class TestExtractMetadata:
    """Tests for transaction metadata."""

    def test_transaction_has_type_metadata(self, basic_importer, minimal_excel_file):
        """Transaction includes type metadata."""
        entries = basic_importer.extract(str(minimal_excel_file), [])

        transactions = [e for e in entries if isinstance(e, data.Transaction)]
        txn = transactions[0]
        assert "type" in txn.meta
        assert txn.meta["type"] == "DEBIT"

    def test_transaction_has_file_location_metadata(
        self, basic_importer, minimal_excel_file
    ):
        """Transaction includes source file location in metadata."""
        entries = basic_importer.extract(str(minimal_excel_file), [])

        transactions = [e for e in entries if isinstance(e, data.Transaction)]
        txn = transactions[0]
        assert "filename" in txn.meta


class TestExtractEdgeCases:
    """Edge cases for extraction."""

    def test_empty_file_returns_empty_list(self, basic_importer, tmp_path):
        """Empty Excel file returns empty list."""
        wb = Workbook()
        ws = wb.active
        headers = ["Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        file_path = tmp_path / "empty.xlsx"
        wb.save(file_path)

        entries = basic_importer.extract(str(file_path), [])
        assert entries == []

    def test_transaction_with_missing_date_is_skipped(self, basic_importer, tmp_path):
        """Transactions without dates are skipped."""
        wb = Workbook()
        ws = wb.active
        headers = ["Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Transaction with no date
        ws.cell(row=2, column=2, value="NO DATE MERCHANT")
        ws.cell(row=2, column=6, value=50.00)

        # Transaction with date
        ws.cell(row=3, column=1, value=datetime.datetime(2025, 10, 24))
        ws.cell(row=3, column=2, value="HAS DATE MERCHANT")
        ws.cell(row=3, column=6, value=100.00)

        file_path = tmp_path / "mixed.xlsx"
        wb.save(file_path)

        entries = basic_importer.extract(str(file_path), [])
        transactions = [e for e in entries if isinstance(e, data.Transaction)]

        # Only the transaction with a date should be extracted
        assert len(transactions) == 1
        assert transactions[0].narration == "HAS DATE MERCHANT"

    def test_transaction_with_no_amount_is_skipped(self, basic_importer, tmp_path):
        """Transactions without amounts are skipped."""
        wb = Workbook()
        ws = wb.active
        headers = ["Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Transaction with no amount
        ws.cell(row=2, column=1, value=datetime.datetime(2025, 10, 24))
        ws.cell(row=2, column=2, value="NO AMOUNT")

        file_path = tmp_path / "no_amount.xlsx"
        wb.save(file_path)

        entries = basic_importer.extract(str(file_path), [])
        assert entries == []


class TestDateMethod:
    """Tests for the date() method."""

    def test_returns_latest_transaction_date(self, basic_importer, sample_excel_path):
        """date() returns the latest transaction date from the file."""
        if not sample_excel_path.exists():
            pytest.skip("Sample Excel file not available")

        result = basic_importer.date(str(sample_excel_path))
        # The date() method returns the latest date from all transactions
        # including balance forward (2025-11-10)
        assert result == datetime.date(2025, 11, 10)

    def test_returns_today_for_empty_file(self, basic_importer, tmp_path):
        """date() returns today's date for files with no valid transactions."""
        wb = Workbook()
        ws = wb.active
        headers = ["Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        file_path = tmp_path / "empty.xlsx"
        wb.save(file_path)

        result = basic_importer.date(str(file_path))
        assert result == datetime.date.today()
