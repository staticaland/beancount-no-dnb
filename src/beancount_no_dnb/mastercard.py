"""DNB Mastercard Excel importer for Beancount."""

import datetime
import sys
import traceback
import warnings
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

import beangulp
from beancount.core import data
from beancount.core.amount import Amount
from beancount.core.number import D
from beancount_classifier import (
    IMPORT_FINGERPRINT_META_KEY,
    ClassifierMixin,
    ImportFingerprintTracker,
    TransactionPattern,
    entry_import_fingerprint,
)
from beangulp import extract, similar
from openpyxl import load_workbook

from beancount_no_dnb.models import (
    ExcelFileData,
    RawTransaction,
)

# Constants
DEFAULT_CURRENCY = "NOK"

FOREIGN_CURRENCY_META_KEY = "foreign_currency"
EXCHANGE_RATE_META_KEY = "exchange_rate"

# Known description patterns
PAYMENT_DESCRIPTION = "Innbetaling"
BALANCE_FORWARD_DESCRIPTION = "Skyldig beløp fra forrige faktura"

# Expected Excel headers
EXPECTED_HEADERS = ("Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut")


@dataclass
class Config:
    """Configuration for a DNB Mastercard Excel account.

    Attributes:
        account_name: The Beancount account name (e.g., 'Liabilities:CreditCard:DNB')
        currency: Default currency for transactions (e.g., 'NOK')
        transaction_patterns: List of TransactionPattern objects for categorization.
        default_account: Account for unmatched transactions in either direction.
            Shorthand when one fallback is enough.
        default_expense_account: Default account for unmatched expenses
            (amount < 0). Takes precedence over default_account for expenses.
        default_income_account: Default account for unmatched income
            (amount > 0, e.g. refunds). Takes precedence over default_account
            for income.
        default_split_percentage: When set (0-100), matched transactions are split between
            the matched account(s) and default_account. Requires default_account to be set.
        skip_balance_forward: When True, skip "Skyldig beløp fra forrige faktura" entries.
        skip_payments: When True, skip "Innbetaling" entries.
        skip_deduplication: When True, skip import_fingerprint-based deduplication.
        dedup_window_days: Days to look back for duplicates.
        dedup_max_date_delta: Max days difference for duplicate detection.
        dedup_epsilon: Tolerance for amount differences in duplicates.
    """

    account_name: str
    currency: str = DEFAULT_CURRENCY
    transaction_patterns: list[TransactionPattern] = field(default_factory=list)
    default_account: str | None = None
    default_expense_account: str | None = None
    default_income_account: str | None = None
    default_split_percentage: int | float | None = None
    skip_balance_forward: bool = True
    skip_payments: bool = False
    skip_deduplication: bool = False
    dedup_window_days: int = 3
    dedup_max_date_delta: int = 2
    dedup_epsilon: Decimal = Decimal("0.05")


DnbConfig = Config


@dataclass
class DnbMastercardConfig(Config):
    """Deprecated alias for Config."""

    def __post_init__(self) -> None:
        warnings.warn(
            "DnbMastercardConfig is deprecated; use Config or DnbConfig instead.",
            DeprecationWarning,
            stacklevel=2,
        )


def _parse_norwegian_number(value) -> Decimal | None:
    """Parse a number that may use Norwegian format (comma as decimal separator).

    Args:
        value: The value to parse (can be float, int, str, or None)

    Returns:
        A Decimal, or None if the value is empty/None
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    if isinstance(value, str):
        # Handle Norwegian number format: replace comma with period
        cleaned = value.strip().replace(",", ".")
        if not cleaned:
            return None
        return Decimal(cleaned)

    return Decimal(str(value))


def _fingerprint_parts(raw_txn: RawTransaction) -> tuple[str, ...]:
    """Row-content identity parts for DNB Mastercard imports."""
    return (
        str(raw_txn.date or ""),
        raw_txn.description or "",
        raw_txn.foreign_currency or "",
        str(raw_txn.exchange_rate or ""),
        str(raw_txn.credit or ""),
        str(raw_txn.debit or ""),
    )


def _is_dnb_mastercard_file(filepath: str) -> bool:
    """Check if an Excel file is a DNB Mastercard statement.

    Verifies the file has the expected column headers.
    """
    path = Path(filepath)

    if path.suffix.lower() != ".xlsx":
        return False

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Check headers in row 1
        headers = tuple(ws.cell(row=1, column=col).value for col in range(1, 7))
        wb.close()

        return headers == EXPECTED_HEADERS
    except Exception:
        return False


class Importer(ClassifierMixin, beangulp.Importer):
    """Importer for DNB Mastercard Excel statements.

    Inherits transaction classification from ClassifierMixin.

    Note: DNB's Excel export carries no balance information, so unlike the
    Amex and SpareBank 1 importers this one cannot emit balance assertions.
    """

    def __init__(
        self,
        config: Config,
        flag: str = "*",
        debug: bool = False,
    ):
        """Initialize the DNB Mastercard Excel importer.

        Args:
            config: A Config object with account details.
            flag: Transaction flag (default: "*").
            debug: Enable debug output (default: False).
        """
        self.account_name = config.account_name
        self.currency = config.currency
        self.transaction_patterns = config.transaction_patterns
        self.default_account = config.default_account
        self.default_expense = config.default_expense_account
        self.default_income = config.default_income_account
        self.default_split_percentage = (
            Decimal(str(config.default_split_percentage))
            if config.default_split_percentage is not None
            else None
        )
        self.skip_balance_forward = config.skip_balance_forward
        self.skip_payments = config.skip_payments
        self.skip_deduplication = config.skip_deduplication
        self.dedup_window = datetime.timedelta(days=config.dedup_window_days)
        self.dedup_max_date_delta = datetime.timedelta(days=config.dedup_max_date_delta)
        self.dedup_epsilon = config.dedup_epsilon
        self.flag = flag
        self.debug = debug

    def _parse_excel_file(self, filepath: str) -> ExcelFileData:
        """Parse the Excel file and extract transactions."""
        result = ExcelFileData()

        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            result.sheet_name = ws.title

            # Skip header row, process data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                date_val = row[0] if len(row) > 0 else None
                description = row[1] if len(row) > 1 else None
                valuta = row[2] if len(row) > 2 else None
                kurs = row[3] if len(row) > 3 else None
                inn = row[4] if len(row) > 4 else None
                ut = row[5] if len(row) > 5 else None

                # Skip empty rows
                if date_val is None and description is None:
                    continue

                # Convert date if it's a datetime
                txn_date = None
                if date_val is not None:
                    if isinstance(date_val, datetime.datetime):
                        txn_date = date_val.date()
                    elif isinstance(date_val, datetime.date):
                        txn_date = date_val

                raw_txn = RawTransaction(
                    date=txn_date,
                    description=description.strip() if description else None,
                    foreign_currency=valuta.strip() if isinstance(valuta, str) else None,
                    exchange_rate=_parse_norwegian_number(kurs),
                    credit=_parse_norwegian_number(inn),
                    debit=_parse_norwegian_number(ut),
                )

                result.transactions.append(raw_txn)

            wb.close()
            return result

        except Exception:
            if self.debug:
                print(
                    f"Error parsing Excel file: {traceback.format_exc()}",
                    file=sys.stderr,
                )
            return ExcelFileData()

    def identify(self, filepath: str) -> bool:
        """Check if the file is a DNB Mastercard Excel statement."""
        return _is_dnb_mastercard_file(filepath)

    def account(self, filepath: str) -> str:
        """Return the account name for the file."""
        return self.account_name

    def filename(self, filepath: str) -> str:
        """Generate a provider/account/original filename for archived data."""
        base_name = Path(filepath).name
        account_leaf = self.account_name.split(":")[-1]
        return f"dnb.{account_leaf}.{base_name}"

    def date(self, filepath: str) -> datetime.date | None:
        """Extract the latest transaction date from the file."""
        parsed_data = self._parse_excel_file(filepath)

        dates = [
            txn.date
            for txn in parsed_data.transactions
            if txn.date is not None
        ]

        if not dates:
            return None

        return max(dates)

    def extract(
        self, filepath: str, existing_entries: list[data.Directive]
    ) -> list[data.Directive]:
        """Extract transactions from a DNB Mastercard Excel file.

        Args:
            filepath: Path to the Excel file
            existing_entries: Existing directives from the ledger, used for deduplication

        Returns:
            List of extracted Beancount Transaction directives
        """
        entries = []

        # Parse the Excel file
        excel_data = self._parse_excel_file(filepath)
        if not excel_data.transactions:
            if self.debug:
                print(f"No transactions found in {filepath}", file=sys.stderr)
            return []

        fingerprint_tracker = ImportFingerprintTracker()

        # Process each transaction
        for idx, raw_txn in enumerate(excel_data.transactions, 1):
            try:
                # Skip transactions without date
                if raw_txn.date is None:
                    if self.debug:
                        print(
                            f"Skipping transaction {idx}: missing date",
                            file=sys.stderr,
                        )
                    continue

                # Skip balance forward entries if configured
                description = raw_txn.description or ""
                if self.skip_balance_forward and description == BALANCE_FORWARD_DESCRIPTION:
                    if self.debug:
                        print(
                            f"Skipping balance forward entry at row {idx}",
                            file=sys.stderr,
                        )
                    continue

                # Skip payment entries if configured
                if self.skip_payments and description == PAYMENT_DESCRIPTION:
                    if self.debug:
                        print(
                            f"Skipping payment entry at row {idx}",
                            file=sys.stderr,
                        )
                    continue

                # Calculate amount: credits are positive (Inn), debits are negative (Ut)
                if raw_txn.credit is not None:
                    amount_decimal = raw_txn.credit
                elif raw_txn.debit is not None:
                    amount_decimal = -raw_txn.debit
                else:
                    if self.debug:
                        print(
                            f"Skipping transaction {idx}: no amount",
                            file=sys.stderr,
                        )
                    continue

                # Create metadata
                metadata = data.new_metadata(filepath, idx)

                # Add transaction type
                if raw_txn.credit is not None:
                    metadata["type"] = "CREDIT"
                else:
                    metadata["type"] = "DEBIT"

                # Preserve foreign-currency statement data without changing
                # the authoritative NOK posting amount.
                if raw_txn.foreign_currency is not None:
                    metadata[FOREIGN_CURRENCY_META_KEY] = raw_txn.foreign_currency
                if raw_txn.exchange_rate is not None:
                    metadata[EXCHANGE_RATE_META_KEY] = raw_txn.exchange_rate

                # Add deterministic identity for re-import matching
                metadata[IMPORT_FINGERPRINT_META_KEY] = fingerprint_tracker.fingerprint(
                    _fingerprint_parts(raw_txn)
                )

                # Create the primary posting
                amount_obj = Amount(D(str(amount_decimal)), self.currency)
                primary_posting = data.Posting(
                    self.account_name, amount_obj, None, None, None, None
                )

                # Create the transaction
                txn = data.Transaction(
                    meta=metadata,
                    date=raw_txn.date,
                    flag=self.flag,
                    payee=None,
                    narration=description,
                    tags=data.EMPTY_SET,
                    links=data.EMPTY_SET,
                    postings=[primary_posting],
                )

                # Apply classification (adds balancing posting)
                finalized_txn = self.finalize(txn, raw_txn)

                if finalized_txn is None:
                    if self.debug:
                        print(
                            f"Skipping transaction {idx} after finalization",
                            file=sys.stderr,
                        )
                    continue

                entries.append(finalized_txn)

            except Exception as e:
                if self.debug:
                    print(
                        f"Error processing transaction {idx}: {e}\n{traceback.format_exc()}",
                        file=sys.stderr,
                    )
                continue

        if existing_entries:
            self.deduplicate(entries, existing_entries)

        return entries

    def deduplicate(
        self, entries: list[data.Directive], existing: list[data.Directive]
    ) -> None:
        """Mark duplicate entries based on configurable parameters."""
        if self.skip_deduplication:
            return

        heuristic_comparator = similar.heuristic_comparator(
            max_date_delta=self.dedup_max_date_delta,
            epsilon=self.dedup_epsilon,
        )

        def comparator(entry: data.Directive, target: data.Directive) -> bool:
            entry_fingerprint = entry_import_fingerprint(entry)
            target_fingerprint = entry_import_fingerprint(target)
            if entry_fingerprint and target_fingerprint:
                return entry_fingerprint == target_fingerprint
            return heuristic_comparator(entry, target)

        extract.mark_duplicate_entries(entries, existing, self.dedup_window, comparator)
