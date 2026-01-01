"""DNB Mastercard Excel importer for Beancount."""

import datetime
import traceback
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

import beangulp
from beangulp import Ingest
from beangulp.testing import main as test_main
from beancount.core import data
from beancount.core.amount import Amount
from beancount.core.number import D
from openpyxl import load_workbook

from beancount_classifier import (
    AccountSplit,
    ClassifierMixin,
    TransactionPattern,
)
from beancount_no_dnb.models import (
    ExcelFileData,
    ParsedTransaction,
    RawTransaction,
)

# Constants
DEFAULT_CURRENCY = "NOK"

# Known description patterns
PAYMENT_DESCRIPTION = "Innbetaling"
BALANCE_FORWARD_DESCRIPTION = "Skyldig beløp fra forrige faktura"

# Expected Excel headers
EXPECTED_HEADERS = ("Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut")


@dataclass
class DnbMastercardConfig:
    """Configuration for a DNB Mastercard Excel account.

    Attributes:
        account_name: The Beancount account name (e.g., 'Liabilities:CreditCard:DNB')
        currency: Default currency for transactions (e.g., 'NOK')
        transaction_patterns: List of TransactionPattern objects for categorization.
        default_account: Account for unmatched transactions.
        skip_balance_forward: When True, skip "Skyldig beløp fra forrige faktura" entries.
        skip_payments: When True, skip "Innbetaling" entries.
    """

    account_name: str
    currency: str = DEFAULT_CURRENCY
    transaction_patterns: list[TransactionPattern] = field(default_factory=list)
    default_account: str | None = None
    default_split_percentage: int | float | None = None
    skip_balance_forward: bool = True
    skip_payments: bool = False


def _parse_norwegian_number(value) -> Decimal | None:
    """Parse a number that may use Norwegian format (comma as decimal separator).

    Args:
        value: The value to parse (can be float, int, str, or None)

    Returns:
        A Decimal, or None if the value is empty/None
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    if isinstance(value, str):
        # Handle Norwegian number format: replace comma with period
        cleaned = value.strip().replace(",", ".")
        if not cleaned:
            return None
        return Decimal(cleaned)

    return Decimal(str(value))


def _is_dnb_mastercard_file(filepath: str) -> bool:
    """Check if an Excel file is a DNB Mastercard statement.

    Verifies the file has the expected column headers.
    """
    path = Path(filepath)

    if path.suffix.lower() != ".xlsx":
        return False

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Check headers in row 1
        headers = tuple(ws.cell(row=1, column=col).value for col in range(1, 7))
        wb.close()

        return headers == EXPECTED_HEADERS
    except Exception:
        return False


class Importer(ClassifierMixin, beangulp.Importer):
    """Importer for DNB Mastercard Excel statements.

    Inherits transaction classification from ClassifierMixin.
    """

    def __init__(
        self,
        config: DnbMastercardConfig,
        flag: str = "*",
        debug: bool = True,
    ):
        """Initialize the DNB Mastercard Excel importer.

        Args:
            config: A DnbMastercardConfig object with account details.
            flag: Transaction flag (default: "*").
            debug: Enable debug output (default: True).
        """
        self.account_name = config.account_name
        self.currency = config.currency
        self.transaction_patterns = config.transaction_patterns
        self.default_account = config.default_account
        self.default_split_percentage = (
            Decimal(str(config.default_split_percentage))
            if config.default_split_percentage is not None
            else None
        )
        self.skip_balance_forward = config.skip_balance_forward
        self.skip_payments = config.skip_payments
        self.flag = flag
        self.debug = debug

    def _parse_excel_file(self, filepath: str) -> ExcelFileData:
        """Parse the Excel file and extract transactions."""
        result = ExcelFileData()

        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            result.sheet_name = ws.title

            # Skip header row, process data rows
            for row_num in range(2, ws.max_row + 1):
                date_val = ws.cell(row=row_num, column=1).value
                description = ws.cell(row=row_num, column=2).value
                valuta = ws.cell(row=row_num, column=3).value
                kurs = ws.cell(row=row_num, column=4).value
                inn = ws.cell(row=row_num, column=5).value
                ut = ws.cell(row=row_num, column=6).value

                # Skip empty rows
                if date_val is None and description is None:
                    continue

                # Convert date if it's a datetime
                txn_date = None
                if date_val is not None:
                    if isinstance(date_val, datetime.datetime):
                        txn_date = date_val.date()
                    elif isinstance(date_val, datetime.date):
                        txn_date = date_val

                raw_txn = RawTransaction(
                    date=txn_date,
                    description=description.strip() if description else None,
                    foreign_currency=valuta.strip() if isinstance(valuta, str) else None,
                    exchange_rate=_parse_norwegian_number(kurs),
                    credit=_parse_norwegian_number(inn),
                    debit=_parse_norwegian_number(ut),
                )

                result.transactions.append(raw_txn)

            wb.close()
            return result

        except Exception as e:
            if self.debug:
                print(f"Error parsing Excel file: {traceback.format_exc()}")
            return ExcelFileData()

    def identify(self, filepath: str) -> bool:
        """Check if the file is a DNB Mastercard Excel statement."""
        return _is_dnb_mastercard_file(filepath)

    def account(self, filepath: str) -> str:
        """Return the account name for the file."""
        return self.account_name

    def filename(self, filepath: str) -> str:
        """Generate a descriptive filename for the imported data."""
        base_name = Path(filepath).name
        account_suffix = self.account_name.split(":")[-1]
        return f"dnb_mastercard.{account_suffix}.{base_name}"

    def date(self, filepath: str) -> datetime.date | None:
        """Extract the latest transaction date from the file."""
        parsed_data = self._parse_excel_file(filepath)

        dates = [
            txn.date
            for txn in parsed_data.transactions
            if txn.date is not None
        ]

        if not dates:
            return datetime.date.today()

        return max(dates)

    def extract(
        self, filepath: str, existing_entries: list[data.Directive]
    ) -> list[data.Directive]:
        """Extract transactions from a DNB Mastercard Excel file.

        Args:
            filepath: Path to the Excel file
            existing_entries: Existing directives from the ledger (unused for now)

        Returns:
            List of extracted Beancount Transaction directives
        """
        entries = []

        # Parse the Excel file
        excel_data = self._parse_excel_file(filepath)
        if not excel_data.transactions:
            if self.debug:
                print(f"No transactions found in {filepath}")
            return []

        # Process each transaction
        for idx, raw_txn in enumerate(excel_data.transactions, 1):
            try:
                # Skip transactions without date
                if raw_txn.date is None:
                    if self.debug:
                        print(f"Skipping transaction {idx}: missing date")
                    continue

                # Skip balance forward entries if configured
                description = raw_txn.description or ""
                if self.skip_balance_forward and description == BALANCE_FORWARD_DESCRIPTION:
                    if self.debug:
                        print(f"Skipping balance forward entry at row {idx}")
                    continue

                # Skip payment entries if configured
                if self.skip_payments and description == PAYMENT_DESCRIPTION:
                    if self.debug:
                        print(f"Skipping payment entry at row {idx}")
                    continue

                # Calculate amount: credits are positive (Inn), debits are negative (Ut)
                if raw_txn.credit is not None:
                    amount_decimal = raw_txn.credit
                elif raw_txn.debit is not None:
                    amount_decimal = -raw_txn.debit
                else:
                    if self.debug:
                        print(f"Skipping transaction {idx}: no amount")
                    continue

                # Create metadata
                metadata = data.new_metadata(filepath, idx)

                # Add transaction type
                if raw_txn.credit is not None:
                    metadata["type"] = "CREDIT"
                else:
                    metadata["type"] = "DEBIT"

                # Create the primary posting
                amount_obj = Amount(D(str(amount_decimal)), self.currency)
                primary_posting = data.Posting(
                    self.account_name, amount_obj, None, None, None, None
                )

                # Create the transaction
                txn = data.Transaction(
                    meta=metadata,
                    date=raw_txn.date,
                    flag=self.flag,
                    payee=None,
                    narration=description,
                    tags=data.EMPTY_SET,
                    links=data.EMPTY_SET,
                    postings=[primary_posting],
                )

                # Apply classification (adds balancing posting)
                finalized_txn = self.finalize(txn, raw_txn)

                if finalized_txn is None:
                    if self.debug:
                        print(f"Skipping transaction {idx} after finalization")
                    continue

                entries.append(finalized_txn)

            except Exception as e:
                if self.debug:
                    print(f"Error processing transaction {idx}: {e}\n{traceback.format_exc()}")
                continue

        return entries


def get_importers() -> list[beangulp.Importer]:
    """Create and return a list of configured importers."""
    return [
        Importer(
            DnbMastercardConfig(
                account_name="Liabilities:CreditCard:DNB",
                currency="NOK",
                transaction_patterns=[],
            )
        ),
    ]


def main():
    """Entry point for the command-line interface."""
    importers = get_importers()
    ingest = Ingest(importers)
    ingest.main()


def test_main_single():
    """Alternative entry point for single-importer testing."""
    importers = get_importers()
    if importers:
        test_main(importers[0])


if __name__ == "__main__":
    main()
